from __future__ import annotations

from copy import copy
from dataclasses import asdict, dataclass, field
from datetime import datetime
from io import BytesIO, StringIO
import json
from pathlib import Path
import re
from uuid import uuid4

import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


APP_TITLE = "Montador de Tabelas"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)
HEADER_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="F0FDFA")
THIN_BORDER = Border(
    left=Side(style="thin", color="D5DADD"),
    right=Side(style="thin", color="D5DADD"),
    top=Side(style="thin", color="D5DADD"),
    bottom=Side(style="thin", color="D5DADD"),
)


class AppError(RuntimeError):
    pass


@dataclass
class ColumnMapping:
    source_name: str = ""
    target_name: str = ""


@dataclass
class SourceConfig:
    id: str = field(default_factory=lambda: uuid4().hex)
    name: str = "Nova tabela"
    url: str = ""
    extraction_mode: str = "auto"
    table_selector: str = ""
    column_mapping: list[ColumnMapping] = field(default_factory=lambda: [ColumnMapping(), ColumnMapping()])
    table_options: list[dict] = field(default_factory=list)


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ")
    text = re.sub(r"[\ue000-\uf8ff]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_key(value) -> str:
    return normalize_text(value).casefold()


def normalize_workbook_name(raw_name: str) -> str:
    name = normalize_text(raw_name) or "tabelas_projeto"
    if not Path(name).suffix:
        name = f"{name}.xlsx"
    suffix = Path(name).suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise AppError("Use um nome de arquivo com extensão .xlsx, .xlsm, .xltx ou .xltm.")
    return Path(name).name


def clean_dataframe(frame: pd.DataFrame) -> pd.DataFrame:
    df = frame.copy()
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [
            " ".join(normalize_text(part) for part in column if normalize_text(part))
            for column in df.columns
        ]
    else:
        df.columns = [normalize_text(column) for column in df.columns]
    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all").fillna("")
    df = df.map(normalize_text)
    if should_promote_first_row(df):
        df.columns = [normalize_text(value) for value in df.iloc[0].tolist()]
        df = df.iloc[1:].reset_index(drop=True)
    return df.reset_index(drop=True)


def should_promote_first_row(frame: pd.DataFrame) -> bool:
    if frame.empty:
        return False
    columns = [normalize_text(col) for col in frame.columns.tolist()]
    numeric_like = all(column.isdigit() for column in columns)
    unnamed_like = all(not column or column.lower().startswith("unnamed") for column in columns)
    if not (numeric_like or unnamed_like):
        return False
    first_row = [normalize_text(value) for value in frame.iloc[0].tolist()]
    non_empty = [value for value in first_row if value]
    return len(non_empty) >= max(1, len(first_row) - 1)


def parse_html_tables(html: str) -> list[pd.DataFrame]:
    soup = BeautifulSoup(html, "html.parser")
    frames: list[pd.DataFrame] = []
    for table in soup.find_all("table"):
        try:
            parsed = pd.read_html(StringIO(str(table)))
        except ValueError:
            continue
        for frame in parsed:
            cleaned = clean_dataframe(frame)
            if not cleaned.empty and len(cleaned.columns):
                frames.append(cleaned)
    return frames


def fetch_html(source: SourceConfig, logger) -> tuple[str, str]:
    modes = ["html_table", "browser_fallback"] if source.extraction_mode == "auto" else [source.extraction_mode]
    errors = []
    for mode in modes:
        try:
            logger(f"Tentando modo {mode} para {source.url}")
            if mode == "html_table":
                response = requests.get(source.url, headers={"User-Agent": USER_AGENT}, timeout=45)
                response.raise_for_status()
                response.encoding = response.apparent_encoding or response.encoding
                return response.text, mode
            raise AppError("Esse link precisa de tratamento específico, mas ainda não foi mapeado.")
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{mode}: {exc}")
    raise AppError(" | ".join(errors) if errors else "Não foi possível carregar a página.")


def detect_api_tables(source: SourceConfig, logger):
    if "siefreceitas.receita.economia.gov.br" not in source.url:
        return None
    api_url = "https://siefreceitas.receita.economia.gov.br/api/receitas/"
    logger(f"Usando API direta para {source.url}")
    response = requests.get(api_url, headers={"User-Agent": USER_AGENT}, timeout=60)
    response.raise_for_status()
    rows = []
    for item in response.json():
        fundamentos = item.get("fundamentos") or []
        base_legal = []
        detalhe_base = []
        for fundamento in fundamentos:
            tp_ato = ((fundamento.get("tpAto") or {}).get("descricao") or "").strip()
            numero = fundamento.get("numero")
            data_fundamento = fundamento.get("data") or ""
            orgaos = ", ".join(orgao.get("nome", "") for orgao in fundamento.get("orgaos") or [])
            base_legal.append(" ".join(str(part) for part in [tp_ato, numero, data_fundamento] if part))
            if orgaos:
                detalhe_base.append(orgaos)
        rows.append(
            {
                "Código de Receita": normalize_text(item.get("recCd")),
                "Denominação": normalize_text(item.get("recNm")),
                "Base Legal": normalize_text(" | ".join(base_legal)),
                "Detalhe Base Legal": normalize_text(" | ".join(detalhe_base)),
                "Data Criação": normalize_text(item.get("dtInicioVigencia")),
                "Data Extinção": normalize_text(item.get("dtFimVigencia")),
            }
        )
    frame = clean_dataframe(pd.DataFrame(rows))
    logger(f"{len(frame.index)} linhas carregadas pela API.")
    return [frame]


def detect_tables(source: SourceConfig, logger) -> tuple[list[pd.DataFrame], str]:
    if not source.url:
        raise AppError("Informe a URL da fonte.")
    api_frames = detect_api_tables(source, logger)
    if api_frames is not None:
        return api_frames, "api"
    html, resolved_mode = fetch_html(source, logger)
    frames = parse_html_tables(html)
    if not frames:
        raise AppError("Nenhuma tabela foi encontrada nessa página.")
    logger(f"{len(frames)} tabela(s) detectada(s) com {resolved_mode}.")
    return frames, resolved_mode


def table_options(frames: list[pd.DataFrame]) -> list[dict]:
    options = []
    for index, frame in enumerate(frames):
        headers = [normalize_text(column) for column in frame.columns.tolist()]
        options.append(
            {
                "value": str(index),
                "label": f"Tabela {index + 1}: {', '.join(headers[:5])}",
                "headers": headers,
                "row_count": len(frame.index),
            }
        )
    return options


def choose_table(source: SourceConfig, frames: list[pd.DataFrame]) -> pd.DataFrame:
    if source.table_selector.isdigit():
        index = int(source.table_selector)
        if 0 <= index < len(frames):
            return frames[index]
    requested = [normalize_key(item.source_name) for item in source.column_mapping if item.source_name]
    if not requested:
        return frames[0]

    def score(frame: pd.DataFrame) -> int:
        headers = {normalize_key(column) for column in frame.columns}
        return sum(1 for item in requested if item in headers)

    return max(frames, key=score)


def mapped_dataframe(source: SourceConfig, frame: pd.DataFrame) -> pd.DataFrame:
    mappings = [item for item in source.column_mapping if item.source_name.strip()]
    if not mappings:
        return frame.copy()
    available = {normalize_key(column): column for column in frame.columns}
    selected_columns = []
    rename_map = {}
    for mapping in mappings:
        source_column = available.get(normalize_key(mapping.source_name))
        if not source_column:
            raise AppError(f"Coluna não encontrada: {mapping.source_name}")
        selected_columns.append(source_column)
        rename_map[source_column] = mapping.target_name.strip() or mapping.source_name.strip()
    return frame.loc[:, selected_columns].rename(columns=rename_map)


def preview_source(source: SourceConfig, logger, limit: int = 200) -> tuple[pd.DataFrame, str]:
    frames, resolved_mode = detect_tables(source, logger)
    source.table_options = table_options(frames)
    chosen = choose_table(source, frames)
    source.extraction_mode = resolved_mode if source.extraction_mode == "auto" else source.extraction_mode
    source.table_selector = str(frames.index(chosen))
    mapped = mapped_dataframe(source, chosen)
    logger(f"Prévia pronta para '{source.name}' com {len(mapped.index)} linhas.")
    return mapped.head(limit), resolved_mode


def safe_sheet_name(workbook, preferred_name: str) -> str:
    base = (preferred_name or "Tabela").strip()[:31] or "Tabela"
    if base not in workbook.sheetnames:
        return base
    counter = 2
    while True:
        suffix = f"_{counter}"
        candidate = f"{base[:31-len(suffix)]}{suffix}"
        if candidate not in workbook.sheetnames:
            return candidate
        counter += 1


def sanitize_workbook(source_workbook):
    clean_workbook = Workbook()
    clean_workbook.remove(clean_workbook.active)
    for source_sheet in source_workbook.worksheets:
        target_sheet = clean_workbook.create_sheet(source_sheet.title[:31] or "Planilha")
        target_sheet.sheet_view.showGridLines = source_sheet.sheet_view.showGridLines
        target_sheet.freeze_panes = source_sheet.freeze_panes
        for row in source_sheet.iter_rows():
            for cell in row:
                new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
        for key, dimension in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[key].width = dimension.width
        for key, dimension in source_sheet.row_dimensions.items():
            target_sheet.row_dimensions[key].height = dimension.height
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))
    if not clean_workbook.sheetnames:
        worksheet = clean_workbook.create_sheet("Resumo")
        worksheet["A1"] = APP_TITLE
        worksheet["A1"].font = Font(bold=True, size=16)
    return clean_workbook


def ensure_workbook_from_upload(uploaded_bytes: bytes | None):
    if uploaded_bytes:
        return sanitize_workbook(load_workbook(BytesIO(uploaded_bytes)))
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Resumo"
    worksheet["A1"] = APP_TITLE
    worksheet["A2"] = "As novas tabelas serão adicionadas como novas abas."
    worksheet["A1"].font = Font(bold=True, size=16)
    return workbook


def write_sheet(workbook, sheet_name: str, dataframe: pd.DataFrame, source: SourceConfig) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.sheet_view.showGridLines = False
    worksheet["A1"] = source.name
    worksheet["A2"] = source.url
    worksheet["A3"] = f"Exportado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    worksheet["A1"].font = Font(bold=True, size=16)
    worksheet["A2"].font = Font(size=10, italic=True, color="64748B")
    worksheet["A3"].font = Font(size=10, color="64748B")
    start_row = 5
    for col_index, column_name in enumerate(dataframe.columns, start=1):
        cell = worksheet.cell(row=start_row, column=col_index, value=column_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    for row_offset, row in enumerate(dataframe.itertuples(index=False), start=1):
        excel_row = start_row + row_offset
        for col_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=excel_row, column=col_index, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_offset % 2 == 0:
                cell.fill = ALT_FILL
    worksheet.freeze_panes = f"A{start_row + 1}"
    if dataframe.shape[1] > 0:
        end_col = get_column_letter(dataframe.shape[1])
        end_row = max(start_row + len(dataframe.index), start_row)
        worksheet.auto_filter.ref = f"A{start_row}:{end_col}{end_row}"
    for col_index, column_cells in enumerate(worksheet.iter_cols(min_row=start_row, max_row=worksheet.max_row), start=1):
        size = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(col_index)].width = min(max(size + 3, 14), 70)


def source_to_dict(source: SourceConfig) -> dict:
    return asdict(source)


def source_from_dict(payload: dict) -> SourceConfig:
    return SourceConfig(
        id=str(payload.get("id") or uuid4().hex),
        name=str(payload.get("name", "Nova tabela")).strip() or "Nova tabela",
        url=str(payload.get("url", "")).strip(),
        extraction_mode=str(payload.get("extraction_mode", "auto")).strip() or "auto",
        table_selector=str(payload.get("table_selector", "")).strip(),
        column_mapping=[
            ColumnMapping(
                source_name=str(item.get("source_name", "")).strip(),
                target_name=str(item.get("target_name", "")).strip(),
            )
            for item in payload.get("column_mapping", [])
        ] or [ColumnMapping(), ColumnMapping()],
        table_options=[dict(item) for item in payload.get("table_options", []) if isinstance(item, dict)],
    )


def init_state() -> None:
    if "project_name" not in st.session_state:
        st.session_state.project_name = "Projeto de Tabelas"
    if "output_name" not in st.session_state:
        st.session_state.output_name = "tabelas_projeto.xlsx"
    if "sources" not in st.session_state:
        st.session_state.sources = [source_to_dict(SourceConfig())]
    if "frames_cache" not in st.session_state:
        st.session_state.frames_cache = {}
    if "preview_cache" not in st.session_state:
        st.session_state.preview_cache = {}
    if "log_messages" not in st.session_state:
        st.session_state.log_messages = []


def add_log(message: str) -> None:
    st.session_state.log_messages.append(f"{datetime.now().strftime('%H:%M:%S')} | {message}")


def add_source() -> None:
    st.session_state.sources.append(source_to_dict(SourceConfig(name=f"Nova tabela {len(st.session_state.sources) + 1}")))


def remove_source(index: int) -> None:
    source_id = st.session_state.sources[index]["id"]
    st.session_state.sources.pop(index)
    st.session_state.frames_cache.pop(source_id, None)
    st.session_state.preview_cache.pop(source_id, None)
    if not st.session_state.sources:
        st.session_state.sources = [source_to_dict(SourceConfig())]


def analyze_source(index: int) -> None:
    source = source_from_dict(st.session_state.sources[index])
    frames, mode = detect_tables(source, add_log)
    source.table_options = table_options(frames)
    source.extraction_mode = mode if source.extraction_mode == "auto" else source.extraction_mode
    if not source.table_selector:
        source.table_selector = "0"
    st.session_state.frames_cache[source.id] = frames
    st.session_state.preview_cache[source.id] = choose_table(source, frames).head(25)
    st.session_state.sources[index] = source_to_dict(source)
    add_log(f"Análise concluída para {source.name}.")


def update_source_from_widgets(index: int) -> SourceConfig:
    payload = st.session_state.sources[index]
    source = source_from_dict(payload)
    source.name = st.session_state.get(f"name_{source.id}", source.name).strip() or "Nova tabela"
    source.url = st.session_state.get(f"url_{source.id}", source.url).strip()
    source.extraction_mode = st.session_state.get(f"mode_{source.id}", source.extraction_mode)
    selected_option = st.session_state.get(f"table_{source.id}", source.table_selector)
    source.table_selector = str(selected_option).split(" - ", 1)[0] if selected_option else source.table_selector
    mapping_rows = st.session_state.get(f"mapping_{source.id}")
    if mapping_rows is not None:
        cleaned_rows = []
        for row in mapping_rows.to_dict("records"):
            cleaned_rows.append(
                ColumnMapping(
                    source_name=normalize_text(row.get("Coluna de origem", "")),
                    target_name=normalize_text(row.get("Coluna final", "")),
                )
            )
        source.column_mapping = cleaned_rows or [ColumnMapping(), ColumnMapping()]
    st.session_state.sources[index] = source_to_dict(source)
    return source


def preview_source_for_ui(index: int) -> tuple[pd.DataFrame, str]:
    source = update_source_from_widgets(index)
    frame, mode = preview_source(source, add_log, limit=250)
    st.session_state.preview_cache[source.id] = frame
    st.session_state.sources[index] = source_to_dict(source)
    return frame, mode


def build_workbook_bytes(uploaded_bytes: bytes | None) -> bytes:
    workbook = ensure_workbook_from_upload(uploaded_bytes)
    for index, payload in enumerate(st.session_state.sources):
        source = source_from_dict(payload)
        source = update_source_from_widgets(index)
        dataframe, _mode = preview_source(source, add_log, limit=1_000_000)
        sheet_name = safe_sheet_name(workbook, source.name)
        write_sheet(workbook, sheet_name, dataframe, source)
        st.session_state.sources[index] = source_to_dict(source)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def load_project_from_json(uploaded) -> None:
    payload = json.loads(uploaded.getvalue().decode("utf-8"))
    st.session_state.project_name = payload.get("project_name", "Projeto de Tabelas")
    st.session_state.output_name = payload.get("output_name", "tabelas_projeto.xlsx")
    st.session_state.sources = [source_to_dict(source_from_dict(item)) for item in payload.get("sources", [])] or [source_to_dict(SourceConfig())]
    st.session_state.frames_cache = {}
    st.session_state.preview_cache = {}
    st.session_state.log_messages = []
    add_log("Projeto carregado do arquivo JSON.")


def export_project_json() -> bytes:
    payload = {
        "project_name": st.session_state.project_name,
        "output_name": st.session_state.output_name,
        "sources": st.session_state.sources,
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


st.set_page_config(page_title=APP_TITLE, page_icon="📊", layout="wide")
init_state()

st.markdown(
    """
    <style>
    .stApp { background: #f7f4ec; }
    .block-container { padding-top: 1.2rem; padding-bottom: 2rem; }
    .app-card {
        background: white; border: 1px solid #eadfce; border-radius: 18px;
        padding: 1rem 1.2rem; box-shadow: 0 6px 18px rgba(15, 23, 42, 0.04);
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title(APP_TITLE)
st.caption("Versão web em Streamlit para publicar e compartilhar.")

with st.sidebar:
    st.subheader("Projeto")
    st.session_state.project_name = st.text_input("Nome do projeto", value=st.session_state.project_name)
    st.session_state.output_name = st.text_input("Nome do Excel", value=st.session_state.output_name)
    if st.button("Adicionar tabela", use_container_width=True):
        add_source()
        st.rerun()

    uploaded_project = st.file_uploader("Carregar projeto JSON", type=["json"])
    if uploaded_project and st.button("Usar projeto carregado", use_container_width=True):
        load_project_from_json(uploaded_project)
        st.rerun()

    st.download_button(
        "Baixar projeto JSON",
        data=export_project_json(),
        file_name="montador_tabelas_projeto.json",
        mime="application/json",
        use_container_width=True,
    )

    st.divider()
    st.subheader("Workbook base opcional")
    base_workbook = st.file_uploader("Enviar Excel existente", type=["xlsx", "xlsm", "xltx", "xltm"])

    with st.expander("Log"):
        if st.session_state.log_messages:
            st.code("\n".join(st.session_state.log_messages[-20:]), language="text")
        else:
            st.caption("Nenhum evento ainda.")

top_a, top_b, top_c = st.columns([1, 1, 2])
top_a.metric("Fontes", len(st.session_state.sources))
preview_count = sum(len(frame.index) for frame in st.session_state.preview_cache.values())
top_b.metric("Linhas em prévia", preview_count)
top_c.info("Configure as fontes, analise, mapeie e depois baixe um único Excel.")

for index, payload in enumerate(st.session_state.sources):
    source = source_from_dict(payload)
    with st.container(border=True):
        header_cols = st.columns([4, 1])
        header_cols[0].subheader(source.name or f"Tabela {index + 1}")
        if header_cols[1].button("Remover", key=f"remove_{source.id}", use_container_width=True):
            remove_source(index)
            st.rerun()

        form_cols = st.columns(2)
        st.session_state.setdefault(f"name_{source.id}", source.name)
        st.session_state.setdefault(f"url_{source.id}", source.url)
        st.session_state.setdefault(f"mode_{source.id}", source.extraction_mode)
        form_cols[0].text_input("Nome da aba", key=f"name_{source.id}")
        form_cols[1].selectbox("Modo de extração", ["auto", "html_table", "browser_fallback"], key=f"mode_{source.id}")
        st.text_input("URL da fonte", key=f"url_{source.id}")

        action_cols = st.columns([1, 1, 1])
        if action_cols[0].button("Analisar tabelas", key=f"analyze_{source.id}", use_container_width=True):
            try:
                update_source_from_widgets(index)
                analyze_source(index)
                st.rerun()
            except Exception as exc:  # noqa: BLE001
                st.error(str(exc))
        if action_cols[1].button("Gerar prévia", key=f"preview_{source.id}", use_container_width=True):
            try:
                frame, mode = preview_source_for_ui(index)
                add_log(f"Prévia atualizada em {mode} para {source.name}.")
                st.success(f"Prévia pronta com {len(frame.index)} linhas.")
                st.rerun()
            except Exception as exc:  # noqa: BLE001
                st.error(str(exc))

        current_source = source_from_dict(st.session_state.sources[index])
        option_labels = [f"{item['value']} - {item['label']}" for item in current_source.table_options]
        selected_label = next(
            (label for label in option_labels if label.startswith(f"{current_source.table_selector} -")),
            option_labels[0] if option_labels else "",
        )
        if option_labels:
            st.selectbox("Tabela detectada", option_labels, index=option_labels.index(selected_label), key=f"table_{source.id}")

        mapping_seed = pd.DataFrame(
            [
                {
                    "Coluna de origem": item.source_name,
                    "Coluna final": item.target_name,
                }
                for item in current_source.column_mapping
            ]
            or [{"Coluna de origem": "", "Coluna final": ""}]
        )
        edited_mapping = st.data_editor(
            mapping_seed,
            key=f"mapping_{source.id}",
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
        )
        st.session_state[f"mapping_{source.id}"] = edited_mapping

        preview_frame = st.session_state.preview_cache.get(source.id)
        if preview_frame is not None and not preview_frame.empty:
            st.caption(f"Prévia de {min(len(preview_frame.index), 25)} linha(s)")
            st.dataframe(preview_frame, use_container_width=True, hide_index=True)
        elif current_source.table_options:
            st.info("Tabela analisada. Gere a prévia para visualizar os dados já mapeados.")

st.divider()
st.subheader("Exportação")
st.caption("Baixe um único Excel com uma aba para cada fonte configurada.")

download_name = normalize_workbook_name(st.session_state.output_name)
export_cols = st.columns([1, 1])
build_now = export_cols[0].button("Preparar arquivo Excel", use_container_width=True, type="primary")

if build_now:
    try:
        workbook_bytes = build_workbook_bytes(base_workbook.getvalue() if base_workbook else None)
        st.session_state["workbook_bytes"] = workbook_bytes
        add_log("Arquivo Excel consolidado preparado.")
        st.success("Arquivo pronto para download.")
    except Exception as exc:  # noqa: BLE001
        st.error(str(exc))

if st.session_state.get("workbook_bytes"):
    export_cols[1].download_button(
        "Baixar Excel consolidado",
        data=st.session_state["workbook_bytes"],
        file_name=download_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
